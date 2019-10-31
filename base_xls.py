#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/10/1 20:13
# @Author  : Wind
# @Des     : 
# @Site    : 
# @File    : base_xls.py
# @Software: PyCharm

# 读写2003 excel
import xlrd
import xlwt
# 读写2007 excel
import openpyxl


# excel 2003写入
def write_03_excel(data):
    # @ data.path  输入值，文件保存路径
    # @ data.sheet_name  输入值，excel分页名称
    # @ data.value  输入值，要保存的数据
    wb = xlwt.Workbook()
    sheet = wb.add_sheet(data.sheet_name)
    value = data.value
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(data.path)
    print("写入数据成功！")


# excel 2003读取
def read_03_excel(path):
    """
    :param path:文件路径
    :yield 生成器
    """
    # @ path 读取文件的路径
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    # print(sheets)
    worksheet = workbook.sheet_by_name(sheets[0])

    for i in range(0, worksheet.nrows):
        # row = worksheet.row(i)
        data_list_item = []
        for j in range(0, worksheet.ncols):
            # print(worksheet.cell_value(i, j), "\t", end="")
            data_list_item.append(worksheet.cell_value(i, j))
        yield (data_list_item)


# excel 2007写入
def write_07_excel(data):
    """
    :param data:
    """
    # @ data.path  输入值，文件保存路径
    # @ data.sheet_name  输入值，excel分页名称
    # @ data.value  输入值，要保存的数据
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = data.sheet_name

    value = data.value
    for i in range(4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(data.path)
    print("写入数据成功！")


# excel 2007读入
def read_07_excel(path):
    """
    :param path: 读取文件的路径
    :yield: 返回生成器
    """
    # @ path  输入值，
    wb = openpyxl.load_workbook(path)
    # print(wb.worksheets)
    sheet = wb.worksheets[0]

    for row in sheet.rows:
        data_item = []
        for cell in row:
            # print(cell.value, "\t", end="")
            data_item.append(cell.value)
        yield data_item

